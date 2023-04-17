import time
import openpyxl
from selenium import webdriver
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from parserSettings import ALI_INSIDER_PAGE, SHIP_TO_ID, SHIP_TO_ADRESS_CLASS, MAIN_PAGE
from parseProduct import parseProduct
import json
import os
from selenium.webdriver.common.by import By
import shutil

# DELAYS
DELAY_ONLOAD = 9
DELAY_ONLOAD_SPEED = 4
DELAY_ONLOAD_AUTH = 20
DELAY_ONLOAD_DOWNLOAD_ALI_INSIDER = 15
DELAY_ONCLICK = 5
DELAY_TIK = 2

# Запускаем браузер
driver = webdriver.Chrome()


class Product:
    def __init__(self, link, title, imgPath, reviewCount, solds, price, deliveryPrice, delivery, shopName, monthSolds, monthRevenue, isStar):
        self.link = link
        self.title = title
        self.imgPath = imgPath
        self.reviewCount = reviewCount
        self.solds = solds

        # CARD INFO
        self.price = price
        self.deliveryPrice = deliveryPrice
        self.delivery = delivery
        self.shopName = shopName

        # AliInsider
        self.monthSolds = monthSolds
        self.monthRevenue = monthRevenue
        self.isStar = isStar


class ProductCategory:
    def __init__(self, name):
        self.name = name
        self.content = []


driver.get(ALI_INSIDER_PAGE)
time.sleep(DELAY_ONLOAD_DOWNLOAD_ALI_INSIDER)

driver.get(MAIN_PAGE)
time.sleep(DELAY_TIK)


# SET RIGHT SHIP TO INFORMATION
shipToSelectElement = driver.find_element(By.ID, SHIP_TO_ID)
shipToSelectElement.click()
time.sleep(DELAY_ONCLICK)

addressInput = driver.find_elements(By.CLASS_NAME, SHIP_TO_ADRESS_CLASS)[0]
addressInput.click()
time.sleep(DELAY_TIK)


driver.find_elements(By.CLASS_NAME, 'address-select-item')[206].click()
time.sleep(DELAY_TIK)

languageElement = driver.find_elements(
    By.CLASS_NAME, 'switcher-currency-c')[0].find_element(By.TAG_NAME, 'span')
languageElement.click()
time.sleep(DELAY_ONLOAD_SPEED)

addressInput = driver.find_elements(By.CLASS_NAME, 'switcher-item')[0]
addressInput.click()
time.sleep(DELAY_TIK)


saveButton = driver.find_element(By.CLASS_NAME, 'switcher-btn').find_element(
    By.TAG_NAME, 'button')
saveButton.click()
time.sleep(DELAY_ONCLICK)

# Укажите путь к директории, которую нужно прочитать
path = "jsons-1"

# Получите список всех файлов в директории
file_list = os.listdir(path)

for file_name in file_list:
    file_path = os.path.join(path, file_name)

    if os.path.isfile(file_path) and '.json' in file_path:
        # Открываем файл в режиме чтения и читаем его содержимое в строку
        print(file_path)
        with open(file_path, 'r') as file:
            data_str = file.read()

        # Преобразуем содержимое файла в объект Python
        data_obj = json.loads(data_str)
        category = ProductCategory(data_obj["categoryName"])
        for productObj in (data_obj["JSONs"]):
            print(productObj["itemUrl"])
            productReady = parseProduct(driver, productObj)
            if (productReady != False):
                category.content.append(productReady)

        # Создаем Excel-файл и записываем данные в него
        wbSourse = openpyxl.load_workbook('data.xlsx')
        shutil.copy(
            'data.xlsx', f'excel/data-{data_obj["categoryName"]}.xlsx')
        wb = openpyxl.load_workbook(
            f'excel/data-{data_obj["categoryName"]}.xlsx')
        ws = wb.active

        index = 2
        ws.cell(row=index, column=1, value=category.name)
        for content in category.content:
            ws.cell(row=index+1, column=2, value=content["itemUrl"])

            response = requests.get(content["itemMainPic"])
            img = Image(BytesIO(response.content))
            img.width = 320
            img.height = 320
            ws.add_image(img, 'E' + str(index+1))

            ws.cell(row=index+1, column=6, value=content["itemName"])
            ws.cell(row=index+1, column=7,
                    value=content["rangePriceFormat"])
            ws.cell(row=index+1, column=8, value=content["shopName"])
            ws.cell(row=index+1, column=9, value=content["score"])
            ws.cell(row=index+1, column=10, value=content["orders"])
            ws.cell(row=index+1, column=11, value=content["deliveryPrice"])
            ws.cell(row=index+1, column=12, value=content["delivery"])
            ws.cell(row=index+1, column=13, value=content["monthSolds"])
            ws.cell(row=index+1, column=14, value=content["monthRevenue"])
            ws.cell(row=index+1, column=15, value=content["isStar"])
            index += 1
        wb.save(f'excel/data-{data_obj["categoryName"]}.xlsx')

# Закрываем браузер
driver.quit()
